"""
Django Template Views - واجهة الويب
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView, TemplateView
from django.urls import reverse_lazy
from django.contrib import messages
from django.db.models import Count, Sum, Q
from django.http import JsonResponse, HttpResponseForbidden
from functools import wraps

from apps.branches.models import Branch
from apps.banks.models import Bank
from apps.cost_centers.models import CostCenter
from apps.suppliers.models import Supplier
from apps.payments.models import PaymentRequest, PaymentRequestItem
from apps.core.models import Role


# =============================================================================
# Custom Decorators
# =============================================================================

def admin_required(view_func):
    """تحقق من أن المستخدم لديه دور أدمن"""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('auth:login')
        
        # التحقق من دور الأدمن
        try:
            user_role = request.user.profile.role
            if user_role and user_role.role_type == Role.RoleType.ADMIN:
                return view_func(request, *args, **kwargs)
        except:
            pass
        
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    return wrapper


class AdminRequiredMixin:
    """Mixin للتحقق من دور الأدمن في CBVs"""
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('auth:login')
        
        try:
            user_role = request.user.profile.role
            if user_role and user_role.role_type == Role.RoleType.ADMIN:
                return super().dispatch(request, *args, **kwargs)
        except:
            pass
        
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')


# =============================================================================
# Authentication Views
# =============================================================================

def login_view(request):
    """صفحة تسجيل الدخول"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        remember = request.POST.get('remember')
        
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            if not remember:
                request.session.set_expiry(0)
            messages.success(request, f'مرحباً {user.get_full_name() or user.username}')
            return redirect('dashboard')
        else:
            messages.error(request, 'اسم المستخدم أو كلمة المرور غير صحيحة')
    
    return render(request, 'auth/login.html')


def logout_view(request):
    """تسجيل الخروج"""
    if request.method == 'POST' or request.method == 'GET':
        logout(request)
        messages.success(request, 'تم تسجيل الخروج بنجاح')
        return redirect('auth:login')
    return redirect('dashboard')


# =============================================================================
# Dashboard View
# =============================================================================

@login_required
def dashboard_view(request):
    """لوحة التحكم الرئيسية"""
    stats = {
        'total_payments': PaymentRequest.objects.count(),
        'pending_payments': PaymentRequest.objects.filter(status__in=['draft', 'proposed']).count(),
        'approved_payments': PaymentRequest.objects.filter(status='final_approved').count(),
        'total_amount': PaymentRequest.objects.aggregate(total=Sum('amount'))['total'] or 0,
        'total_branches': Branch.objects.filter(is_active=True).count(),
        'total_banks': Bank.objects.filter(is_active=True).count(),
        'total_suppliers': Supplier.objects.filter(is_active=True).count(),
        'total_cost_centers': CostCenter.objects.filter(is_active=True).count(),
    }
    recent_payments = PaymentRequest.objects.select_related('branch', 'bank').order_by('-created_at')[:5]
    
    context = {
        'stats': stats,
        'recent_payments': recent_payments,
    }
    return render(request, 'pages/dashboard.html', context)


# =============================================================================
# Branch Views
# =============================================================================

class BranchListView(LoginRequiredMixin, ListView):
    """قائمة الفروع"""
    model = Branch
    template_name = 'pages/branches/list.html'
    context_object_name = 'object_list'
    paginate_by = 15
    
    def get_queryset(self):
        queryset = Branch.objects.select_related('cost_center')
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(Q(name__icontains=search) | Q(code__icontains=search))
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'الفروع'
        context['page_subtitle'] = 'إدارة فروع المؤسسة'
        context['add_button_text'] = 'إضافة فرع'
        context['search_placeholder'] = 'البحث عن فرع...'
        context['cost_centers'] = CostCenter.objects.filter(is_active=True)
        return context


@login_required
def branch_create(request):
    """إنشاء فرع جديد"""
    if request.method == 'POST':
        Branch.objects.create(
            name=request.POST.get('name'),
            code=request.POST.get('code'),
            cost_center_id=request.POST.get('cost_center') or None,
            address=request.POST.get('address', ''),
            phone=request.POST.get('phone', ''),
        )
        messages.success(request, 'تم إضافة الفرع بنجاح')
    return redirect('branches:list')


@login_required
def branch_update(request, pk):
    """تعديل فرع"""
    branch = get_object_or_404(Branch, pk=pk)
    if request.method == 'POST':
        branch.name = request.POST.get('name')
        branch.code = request.POST.get('code')
        branch.cost_center_id = request.POST.get('cost_center') or None
        branch.address = request.POST.get('address', '')
        branch.phone = request.POST.get('phone', '')
        branch.save()
        messages.success(request, 'تم تعديل الفرع بنجاح')
    return redirect('branches:list')


@login_required
def branch_delete(request, pk):
    """حذف فرع"""
    branch = get_object_or_404(Branch, pk=pk)
    if request.method == 'POST':
        branch.delete()
        messages.success(request, 'تم حذف الفرع بنجاح')
    return redirect('branches:list')


# =============================================================================
# Bank Views
# =============================================================================

class BankListView(LoginRequiredMixin, ListView):
    """قائمة البنوك"""
    model = Bank
    template_name = 'pages/banks/list.html'
    context_object_name = 'object_list'
    paginate_by = 15
    
    def get_queryset(self):
        queryset = Bank.objects.select_related('branch')
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(Q(name__icontains=search) | Q(code__icontains=search))
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'البنوك'
        context['page_subtitle'] = 'إدارة حسابات البنوك'
        context['add_button_text'] = 'إضافة بنك'
        context['search_placeholder'] = 'البحث عن بنك...'
        context['branches'] = Branch.objects.filter(is_active=True)
        return context


@login_required
def bank_create(request):
    """إنشاء بنك جديد"""
    if request.method == 'POST':
        Bank.objects.create(
            name=request.POST.get('name'),
            code=request.POST.get('code'),
            branch_id=request.POST.get('branch') or None,
            account_number=request.POST.get('account_number', ''),
            iban=request.POST.get('iban', ''),
        )
        messages.success(request, 'تم إضافة البنك بنجاح')
    return redirect('banks:list')


@login_required
def bank_update(request, pk):
    """تعديل بنك"""
    bank = get_object_or_404(Bank, pk=pk)
    if request.method == 'POST':
        bank.name = request.POST.get('name')
        bank.code = request.POST.get('code')
        bank.branch_id = request.POST.get('branch') or None
        bank.account_number = request.POST.get('account_number', '')
        bank.iban = request.POST.get('iban', '')
        bank.save()
        messages.success(request, 'تم تعديل البنك بنجاح')
    return redirect('banks:list')


@login_required
def bank_delete(request, pk):
    """حذف بنك"""
    bank = get_object_or_404(Bank, pk=pk)
    if request.method == 'POST':
        bank.delete()
        messages.success(request, 'تم حذف البنك بنجاح')
    return redirect('banks:list')


# =============================================================================
# Cost Center Views
# =============================================================================

class CostCenterListView(LoginRequiredMixin, ListView):
    """قائمة مراكز التكلفة"""
    model = CostCenter
    template_name = 'pages/cost_centers/list.html'
    context_object_name = 'object_list'
    paginate_by = 15
    
    def get_queryset(self):
        queryset = CostCenter.objects.prefetch_related('branches')
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(Q(name__icontains=search) | Q(code__icontains=search))
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'مراكز التكلفة'
        context['page_subtitle'] = 'إدارة مراكز التكلفة'
        context['add_button_text'] = 'إضافة مركز تكلفة'
        context['search_placeholder'] = 'البحث عن مركز تكلفة...'
        return context


@login_required
def cost_center_create(request):
    """إنشاء مركز تكلفة جديد"""
    if request.method == 'POST':
        CostCenter.objects.create(
            name=request.POST.get('name'),
            code=request.POST.get('code'),
            description=request.POST.get('description', ''),
        )
        messages.success(request, 'تم إضافة مركز التكلفة بنجاح')
    return redirect('cost_centers:list')


@login_required
def cost_center_update(request, pk):
    """تعديل مركز تكلفة"""
    cc = get_object_or_404(CostCenter, pk=pk)
    if request.method == 'POST':
        cc.name = request.POST.get('name')
        cc.code = request.POST.get('code')
        cc.description = request.POST.get('description', '')
        cc.save()
        messages.success(request, 'تم تعديل مركز التكلفة بنجاح')
    return redirect('cost_centers:list')


@login_required
def cost_center_delete(request, pk):
    """حذف مركز تكلفة"""
    cc = get_object_or_404(CostCenter, pk=pk)
    if request.method == 'POST':
        cc.delete()
        messages.success(request, 'تم حذف مركز التكلفة بنجاح')
    return redirect('cost_centers:list')


# =============================================================================
# Supplier Views
# =============================================================================

class SupplierListView(LoginRequiredMixin, ListView):
    """قائمة الموردين"""
    model = Supplier
    template_name = 'pages/suppliers/list.html'
    context_object_name = 'object_list'
    paginate_by = 15
    
    def get_queryset(self):
        queryset = Supplier.objects.all()
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(Q(name__icontains=search) | Q(code__icontains=search))
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'الموردين'
        context['page_subtitle'] = 'إدارة بيانات الموردين'
        context['add_button_text'] = 'إضافة مورد'
        context['search_placeholder'] = 'البحث عن مورد...'
        return context


@login_required
def supplier_create(request):
    """إنشاء مورد جديد"""
    if request.method == 'POST':
        Supplier.objects.create(
            name=request.POST.get('name'),
            code=request.POST.get('code'),
            phone=request.POST.get('phone', ''),
            email=request.POST.get('email', ''),
            bank_name=request.POST.get('bank_name', ''),
            bank_account=request.POST.get('bank_account', ''),
            iban=request.POST.get('iban', ''),
        )
        messages.success(request, 'تم إضافة المورد بنجاح')
    return redirect('suppliers:list')


@login_required
def supplier_update(request, pk):
    """تعديل مورد"""
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        supplier.name = request.POST.get('name')
        supplier.code = request.POST.get('code')
        supplier.phone = request.POST.get('phone', '')
        supplier.email = request.POST.get('email', '')
        supplier.bank_name = request.POST.get('bank_name', '')
        supplier.bank_account = request.POST.get('bank_account', '')
        supplier.iban = request.POST.get('iban', '')
        supplier.save()
        messages.success(request, 'تم تعديل المورد بنجاح')
    return redirect('suppliers:list')


@login_required
def supplier_delete(request, pk):
    """حذف مورد"""
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        supplier.delete()
        messages.success(request, 'تم حذف المورد بنجاح')
    return redirect('suppliers:list')


# =============================================================================
# Payment Views
# =============================================================================

class PaymentListView(LoginRequiredMixin, ListView):
    """قائمة طلبات الدفع"""
    model = PaymentRequest
    template_name = 'pages/payments/list.html'
    context_object_name = 'object_list'
    paginate_by = 15
    
    def get_queryset(self):
        queryset = PaymentRequest.objects.select_related('branch', 'bank', 'created_by')
        
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(Q(pk__icontains=search) | Q(notes__icontains=search))
        
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        branch = self.request.GET.get('branch')
        if branch:
            queryset = queryset.filter(branch_id=branch)
        
        return queryset.order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['branches'] = Branch.objects.filter(is_active=True)
        return context


class PaymentDetailView(LoginRequiredMixin, DetailView):
    """تفاصيل طلب الدفع"""
    model = PaymentRequest
    template_name = 'pages/payments/detail.html'
    context_object_name = 'payment'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['can_change_status'] = True  # TODO: Add permission check
        return context


@login_required
def payment_create(request):
    """إنشاء طلب دفع جديد"""
    if request.method == 'POST':
        payment = PaymentRequest.objects.create(
            branch_id=request.POST.get('branch'),
            bank_id=request.POST.get('bank') or None,
            cost_center=request.POST.get('cost_center', ''),
            bank_account_number=request.POST.get('bank_account_number', ''),
            supplier_account=request.POST.get('supplier_account', '21101001'),
            analytical_account=request.POST.get('analytical_account', ''),
            notes=request.POST.get('notes', ''),
            created_by=request.user,
            status='draft' if request.POST.get('action') == 'draft' else 'proposed',
        )
        
        # Add items
        items_count = int(request.POST.get('items_count', 0))
        total = 0
        total_proposed = 0
        for i in range(items_count):
            supplier_id = request.POST.get(f'items-{i}-supplier')
            current_balance = request.POST.get(f'items-{i}-current_balance', 0)
            proposed_amount = request.POST.get(f'items-{i}-proposed_amount', 0)
            sultan_approval = request.POST.get(f'items-{i}-sultan_approval', 'pending')
            auditor_status = request.POST.get(f'items-{i}-auditor_status', 'pending')
            cfo_approval = request.POST.get(f'items-{i}-cfo_approval', 'pending')
            abu_alaa_proposed = request.POST.get(f'items-{i}-abu_alaa_proposed', 0)
            abu_alaa_final = request.POST.get(f'items-{i}-abu_alaa_final', 'pending')
            
            if supplier_id:
                PaymentRequestItem.objects.create(
                    payment_request=payment,
                    supplier_id=supplier_id,
                    current_balance=float(current_balance or 0),
                    proposed_amount=float(proposed_amount or 0),
                    sultan_approval=sultan_approval,
                    auditor_status=auditor_status,
                    cfo_approval=cfo_approval,
                    abu_alaa_proposed=float(abu_alaa_proposed or 0),
                    abu_alaa_final=abu_alaa_final,
                    amount=float(abu_alaa_proposed or 0),
                )
                total += float(abu_alaa_proposed or 0)
                total_proposed += float(proposed_amount or 0)
        
        # Update total amount
        payment.amount = total
        payment.save()
        
        messages.success(request, 'تم إنشاء طلب الدفع بنجاح')
        return redirect('payments:detail', pk=payment.pk)
    
    context = {
        'branches': Branch.objects.filter(is_active=True).select_related('cost_center'),
        'banks': Bank.objects.filter(is_active=True).select_related('branch'),
        'cost_centers': CostCenter.objects.filter(is_active=True),
        'suppliers': Supplier.objects.filter(is_active=True),
    }
    return render(request, 'pages/payments/form.html', context)


@login_required
def payment_update(request, pk):
    """تعديل طلب دفع"""
    payment = get_object_or_404(PaymentRequest, pk=pk)
    
    if payment.status != 'draft':
        messages.error(request, 'لا يمكن تعديل طلب غير مسودة')
        return redirect('payments:detail', pk=pk)
    
    if request.method == 'POST':
        payment.branch_id = request.POST.get('branch')
        payment.bank_id = request.POST.get('bank') or None
        payment.cost_center = request.POST.get('cost_center', '')
        payment.bank_account_number = request.POST.get('bank_account_number', '')
        payment.supplier_account = request.POST.get('supplier_account', '21101001')
        payment.analytical_account = request.POST.get('analytical_account', '')
        payment.notes = request.POST.get('notes', '')
        
        if request.POST.get('action') == 'submit':
            payment.status = 'proposed'
        
        payment.save()
        
        # Update items
        payment.items.all().delete()
        items_count = int(request.POST.get('items_count', 0))
        total = 0
        total_proposed = 0
        for i in range(items_count):
            supplier_id = request.POST.get(f'items-{i}-supplier')
            current_balance = request.POST.get(f'items-{i}-current_balance', 0)
            proposed_amount = request.POST.get(f'items-{i}-proposed_amount', 0)
            sultan_approval = request.POST.get(f'items-{i}-sultan_approval', 'pending')
            auditor_status = request.POST.get(f'items-{i}-auditor_status', 'pending')
            cfo_approval = request.POST.get(f'items-{i}-cfo_approval', 'pending')
            abu_alaa_proposed = request.POST.get(f'items-{i}-abu_alaa_proposed', 0)
            abu_alaa_final = request.POST.get(f'items-{i}-abu_alaa_final', 'pending')
            
            if supplier_id:
                PaymentRequestItem.objects.create(
                    payment_request=payment,
                    supplier_id=supplier_id,
                    current_balance=float(current_balance or 0),
                    proposed_amount=float(proposed_amount or 0),
                    sultan_approval=sultan_approval,
                    auditor_status=auditor_status,
                    cfo_approval=cfo_approval,
                    abu_alaa_proposed=float(abu_alaa_proposed or 0),
                    abu_alaa_final=abu_alaa_final,
                    amount=float(abu_alaa_proposed or 0),
                )
                total += float(abu_alaa_proposed or 0)
                total_proposed += float(proposed_amount or 0)
        
        payment.amount = total
        payment.save()
        
        messages.success(request, 'تم تعديل طلب الدفع بنجاح')
        return redirect('payments:detail', pk=pk)
    
    context = {
        'payment': payment,
        'branches': Branch.objects.filter(is_active=True).select_related('cost_center'),
        'banks': Bank.objects.filter(is_active=True).select_related('branch'),
        'cost_centers': CostCenter.objects.filter(is_active=True),
        'suppliers': Supplier.objects.filter(is_active=True),
    }
    return render(request, 'pages/payments/form.html', context)


@login_required
def payment_change_status(request, pk):
    """تغيير حالة طلب الدفع"""
    payment = get_object_or_404(PaymentRequest, pk=pk)
    
    if request.method == 'POST':
        new_status = request.POST.get('status')
        valid_statuses = ['proposed', 'first_approved', 'audited', 'final_approved', 'rejected']
        
        if new_status in valid_statuses:
            payment.status = new_status
            payment.save()
            messages.success(request, f'تم تغيير الحالة إلى {payment.get_status_display()}')
        else:
            messages.error(request, 'حالة غير صالحة')
    
    return redirect('payments:detail', pk=pk)


@login_required
def payment_export_excel(request, pk):
    """تصدير طلب الدفع إلى Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from datetime import datetime
    
    payment = get_object_or_404(PaymentRequest, pk=pk)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"طلب دفع {payment.id}"
    
    # RTL direction
    ws.sheet_view.rightToLeft = True
    
    # Styles
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    title_font = Font(name='Arial', size=14, bold=True)
    title_alignment = Alignment(horizontal='center', vertical='center')
    
    cell_font = Font(name='Arial', size=11)
    cell_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = f'طلب دفع رقم #{payment.id}'
    title_cell.font = title_font
    title_cell.alignment = title_alignment
    
    # Payment info
    ws['A3'] = 'الفرع:'
    ws['B3'] = payment.branch.name
    ws['D3'] = 'التاريخ:'
    ws['E3'] = payment.created_at.strftime('%Y/%m/%d')
    ws['G3'] = 'الحالة:'
    ws['H3'] = payment.get_status_display()
    
    # Headers
    headers = ['رقم المورد', 'اسم المورد', 'رصيد المورد', 'دفعه المشتريات', 
               'اعتماد سلطان', 'حالة المدقق', 'اعتماد المدير المالي', 'اقتراح السداد', 'اعتماد أبوعلاء']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    row_num = 6
    for item in payment.items.all():
        ws.cell(row=row_num, column=1, value=item.supplier.code).alignment = cell_alignment
        ws.cell(row=row_num, column=2, value=item.supplier.name).alignment = cell_alignment
        ws.cell(row=row_num, column=3, value=float(item.current_balance)).alignment = cell_alignment
        ws.cell(row=row_num, column=4, value=float(item.proposed_amount)).alignment = cell_alignment
        ws.cell(row=row_num, column=5, value=item.get_sultan_approval_display()).alignment = cell_alignment
        ws.cell(row=row_num, column=6, value=item.get_auditor_status_display()).alignment = cell_alignment
        ws.cell(row=row_num, column=7, value=item.get_cfo_approval_display()).alignment = cell_alignment
        ws.cell(row=row_num, column=8, value=float(item.abu_alaa_proposed)).alignment = cell_alignment
        ws.cell(row=row_num, column=9, value=item.get_abu_alaa_final_display()).alignment = cell_alignment
        
        # Apply borders
        for col in range(1, 10):
            ws.cell(row=row_num, column=col).border = border
            ws.cell(row=row_num, column=col).font = cell_font
        
        row_num += 1
    
    # Total row
    total_row = row_num
    ws.merge_cells(f'A{total_row}:C{total_row}')
    total_cell = ws.cell(row=total_row, column=1)
    total_cell.value = 'المجموع الكلي'
    total_cell.font = Font(name='Arial', size=12, bold=True)
    total_cell.alignment = Alignment(horizontal='center', vertical='center')
    total_cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    ws.cell(row=total_row, column=4, value=float(payment.total_proposed_amount)).font = Font(bold=True)
    ws.cell(row=total_row, column=4).alignment = cell_alignment
    ws.cell(row=total_row, column=8, value=float(payment.total_abu_alaa_proposed)).font = Font(bold=True)
    ws.cell(row=total_row, column=8).alignment = cell_alignment
    
    # Apply borders to total row
    for col in range(1, 10):
        ws.cell(row=total_row, column=col).border = border
        if col in [1, 4, 8]:
            ws.cell(row=total_row, column=col).fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="payment_{payment.id}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response


# =============================================================================
# User Management Views
# =============================================================================

from django.contrib.auth.models import User

class UserListView(AdminRequiredMixin, LoginRequiredMixin, ListView):
    """قائمة المستخدمين"""
    model = User
    template_name = 'pages/users/list.html'
    context_object_name = 'users'
    
    def get_queryset(self):
        return User.objects.all().order_by('-date_joined')


@login_required
@admin_required
def user_create(request):
    """إنشاء مستخدم جديد"""
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email', '')
        password = request.POST.get('password')
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        is_staff = request.POST.get('is_staff') == 'on'
        is_active = request.POST.get('is_active') == 'on'
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'اسم المستخدم موجود مسبقاً')
        else:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name,
            )
            user.is_staff = is_staff
            user.is_active = is_active
            user.save()
            messages.success(request, 'تم إنشاء المستخدم بنجاح')
            return redirect('users:list')
    
    return render(request, 'pages/users/form.html')


@login_required
@admin_required
def user_update(request, pk):
    """تعديل مستخدم"""
    user = get_object_or_404(User, pk=pk)
    
    if request.method == 'POST':
        user.username = request.POST.get('username')
        user.email = request.POST.get('email', '')
        user.first_name = request.POST.get('first_name', '')
        user.last_name = request.POST.get('last_name', '')
        user.is_staff = request.POST.get('is_staff') == 'on'
        user.is_active = request.POST.get('is_active') == 'on'
        
        password = request.POST.get('password')
        if password:
            user.set_password(password)
        
        user.save()
        messages.success(request, 'تم تعديل المستخدم بنجاح')
        return redirect('users:list')
    
    return render(request, 'pages/users/form.html', {'user_obj': user})


@login_required
@admin_required
def user_delete(request, pk):
    """حذف مستخدم"""
    user = get_object_or_404(User, pk=pk)
    
    if user == request.user:
        messages.error(request, 'لا يمكنك حذف حسابك الخاص')
    else:
        user.delete()
        messages.success(request, 'تم حذف المستخدم بنجاح')
    
    return redirect('users:list')
