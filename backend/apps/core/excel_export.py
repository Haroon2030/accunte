"""
Modern Excel export helpers for web reports.
"""

from datetime import datetime

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ModernExcelBuilder:
    """Build styled RTL Excel workbooks with a consistent modern look."""

    PRIMARY = '2563EB'
    PRIMARY_DARK = '1E40AF'
    HEADER_BG = '2563EB'
    TITLE_BG = 'EFF6FF'
    META_LABEL_BG = 'F1F5F9'
    ALT_ROW = 'F8FAFC'
    TOTAL_BG = 'E2E8F0'
    BORDER_COLOR = 'CBD5E1'

    def __init__(self, report_title, sheet_name='البيانات'):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = sheet_name[:31]
        self.ws.sheet_view.rightToLeft = True
        self.report_title = report_title
        self._styles = self._build_styles()

    def _build_styles(self):
        thin = Side(style='thin', color=self.BORDER_COLOR)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        return {
            'border': border,
            'title_font': Font(name='Segoe UI', size=16, bold=True, color=self.PRIMARY_DARK),
            'subtitle_font': Font(name='Segoe UI', size=10, color='64748B'),
            'meta_label_font': Font(name='Segoe UI', size=10, bold=True, color='334155'),
            'meta_value_font': Font(name='Segoe UI', size=10, color='475569'),
            'header_font': Font(name='Segoe UI', size=11, bold=True, color='FFFFFF'),
            'header_fill': PatternFill(start_color=self.HEADER_BG, end_color=self.HEADER_BG, fill_type='solid'),
            'cell_font': Font(name='Segoe UI', size=10, color='334155'),
            'alt_fill': PatternFill(start_color=self.ALT_ROW, end_color=self.ALT_ROW, fill_type='solid'),
            'total_font': Font(name='Segoe UI', size=11, bold=True, color='0F172A'),
            'total_fill': PatternFill(start_color=self.TOTAL_BG, end_color=self.TOTAL_BG, fill_type='solid'),
            'center': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'right': Alignment(horizontal='right', vertical='center', wrap_text=True),
        }

    def _apply_border(self, cell):
        cell.border = self._styles['border']

    def write_report_header(self, num_columns, meta_items=None):
        """Write title band and optional metadata rows. Returns first data row index."""
        meta_items = meta_items or []
        last_col = get_column_letter(num_columns)

        self.ws.merge_cells(f'A1:{last_col}1')
        title_cell = self.ws['A1']
        title_cell.value = self.report_title
        title_cell.font = self._styles['title_font']
        title_cell.alignment = self._styles['center']
        title_cell.fill = PatternFill(start_color=self.TITLE_BG, end_color=self.TITLE_BG, fill_type='solid')
        self.ws.row_dimensions[1].height = 32

        self.ws.merge_cells(f'A2:{last_col}2')
        subtitle = self.ws['A2']
        subtitle.value = f'تاريخ التصدير: {datetime.now().strftime("%Y/%m/%d %H:%M")}'
        subtitle.font = self._styles['subtitle_font']
        subtitle.alignment = self._styles['center']
        self.ws.row_dimensions[2].height = 20

        row = 4
        if meta_items:
            for label, value in meta_items:
                self.ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                self.ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=num_columns)
                label_cell = self.ws.cell(row=row, column=1, value=label)
                value_cell = self.ws.cell(row=row, column=3, value=value)
                label_cell.font = self._styles['meta_label_font']
                value_cell.font = self._styles['meta_value_font']
                label_cell.alignment = self._styles['right']
                value_cell.alignment = self._styles['right']
                label_cell.fill = PatternFill(start_color=self.META_LABEL_BG, end_color=self.META_LABEL_BG, fill_type='solid')
                for col in range(1, num_columns + 1):
                    self._apply_border(self.ws.cell(row=row, column=col))
                row += 1
            row += 1
        return row

    def write_table(self, start_row, headers, rows, col_widths=None, totals_row=None):
        """Write headers, data rows, and optional totals row."""
        num_cols = len(headers)

        for col_idx, header in enumerate(headers, 1):
            cell = self.ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = self._styles['header_font']
            cell.fill = self._styles['header_fill']
            cell.alignment = self._styles['center']
            self._apply_border(cell)
        self.ws.row_dimensions[start_row].height = 24

        data_row = start_row + 1
        for row_idx, row_values in enumerate(rows):
            current_row = data_row + row_idx
            for col_idx, value in enumerate(row_values, 1):
                cell = self.ws.cell(row=current_row, column=col_idx, value=value)
                cell.font = self._styles['cell_font']
                cell.alignment = self._styles['right'] if col_idx > 1 else self._styles['center']
                self._apply_border(cell)
                if row_idx % 2 == 1:
                    cell.fill = self._styles['alt_fill']
            self.ws.row_dimensions[current_row].height = 20

        if totals_row is not None:
            total_row_num = data_row + len(rows)
            for col_idx, value in enumerate(totals_row, 1):
                cell = self.ws.cell(row=total_row_num, column=col_idx, value=value)
                cell.font = self._styles['total_font']
                cell.fill = self._styles['total_fill']
                cell.alignment = self._styles['right'] if col_idx > 1 else self._styles['center']
                self._apply_border(cell)
            self.ws.row_dimensions[total_row_num].height = 22

        if col_widths:
            for idx, width in enumerate(col_widths, 1):
                self.ws.column_dimensions[get_column_letter(idx)].width = width

        self.ws.freeze_panes = self.ws.cell(row=start_row + 1, column=1)

    def add_sheet(self, sheet_name):
        ws = self.wb.create_sheet(title=sheet_name[:31])
        ws.sheet_view.rightToLeft = True
        return ws

    def write_sheet_table(self, ws, report_title, headers, rows, col_widths=None, meta_items=None, totals_row=None):
        """Write a full styled table on a secondary worksheet."""
        builder = ModernExcelBuilder(report_title, sheet_name=ws.title)
        builder.ws = ws
        builder.wb = self.wb
        start_row = builder.write_report_header(len(headers), meta_items)
        builder.write_table(start_row, headers, rows, col_widths, totals_row)

    def to_response(self, filename_prefix):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        stamp = datetime.now().strftime('%Y%m%d')
        response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{stamp}.xlsx"'
        self.wb.save(response)
        return response
